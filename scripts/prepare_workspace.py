# -*- coding: utf-8 -*-
"""Prepare intermediate workspace from source Excel + external profile."""

from __future__ import annotations

import argparse
import shutil
from datetime import datetime
from pathlib import Path

import openpyxl
import pandas as pd
import yaml


def _indicator_rows_from_profile(profile: dict) -> list[dict]:
    rows = []
    report_month = profile.get("report_month")
    for indicator in profile.get("indicators", []) or []:
        row = dict(indicator)
        if report_month and not row.get("report_month"):
            row["report_month"] = report_month
        if isinstance(row.get("categories"), list):
            row["categories"] = ",".join(row["categories"])
        if isinstance(row.get("filter_exclude"), list):
            row["filter_exclude"] = ",".join(row["filter_exclude"])
        if isinstance(row.get("merge_rules"), dict):
            row["merge_rules"] = ";".join(f"{k}->{v}" for k, v in row["merge_rules"].items())
        rows.append(row)
    return rows


def _styles_rows_from_profile(profile: dict) -> list[dict]:
    styles = profile.get("styles", {}) or {}
    colors = styles.get("colors", {}) or {}
    global_styles = styles.get("global", {}) or {}

    rows = []
    for k, v in colors.items():
        rows.append({"style_key": k, "hex_color": v, "role": "color"})
    for k, v in global_styles.items():
        rows.append({"style_key": k, "hex_color": v, "role": None})
    return rows


def prepare_workspace(
    excel_path: str,
    profile_path: str,
    workspace_root: str = "Data/intermediate/workspaces",
    run_id: str | None = None,
) -> dict:
    src = Path(excel_path)
    profile_file = Path(profile_path)
    if not src.exists():
        raise FileNotFoundError(f"Source Excel not found: {src}")
    if not profile_file.exists():
        raise FileNotFoundError(f"Profile not found: {profile_file}")

    with profile_file.open("r", encoding="utf-8") as f:
        profile = yaml.safe_load(f) or {}

    manual_csv_path = Path(profile.get("manual_inputs_csv", "config/manual_inputs.csv"))
    if not manual_csv_path.exists():
        raise FileNotFoundError(f"manual_inputs_csv not found: {manual_csv_path}")

    if run_id is None:
        run_id = datetime.now().strftime("run_%Y%m%d_%H%M%S")

    workspace = Path(workspace_root) / run_id
    csv_dir = workspace / "csv"
    logs_dir = workspace / "logs"
    workspace.mkdir(parents=True, exist_ok=True)
    csv_dir.mkdir(parents=True, exist_ok=True)
    logs_dir.mkdir(parents=True, exist_ok=True)

    prepared_excel = workspace / "prepared.xlsx"
    shutil.copy2(src, prepared_excel)

    meta_df = pd.DataFrame(_indicator_rows_from_profile(profile))
    styles_df = pd.DataFrame(_styles_rows_from_profile(profile))
    manual_df = pd.read_csv(manual_csv_path)

    wb = openpyxl.load_workbook(prepared_excel)
    for sheet_name in ["_meta", "_styles", "_manual_inputs"]:
        if sheet_name in wb.sheetnames:
            del wb[sheet_name]

    ws_meta = wb.create_sheet("_meta", 0)
    for r_idx, row in enumerate([meta_df.columns.tolist()] + meta_df.values.tolist(), 1):
        for c_idx, value in enumerate(row, 1):
            ws_meta.cell(row=r_idx, column=c_idx, value=value)

    ws_styles = wb.create_sheet("_styles", 1)
    for r_idx, row in enumerate([styles_df.columns.tolist()] + styles_df.values.tolist(), 1):
        for c_idx, value in enumerate(row, 1):
            ws_styles.cell(row=r_idx, column=c_idx, value=value)

    ws_manual = wb.create_sheet("_manual_inputs", 2)
    for r_idx, row in enumerate([manual_df.columns.tolist()] + manual_df.values.tolist(), 1):
        for c_idx, value in enumerate(row, 1):
            ws_manual.cell(row=r_idx, column=c_idx, value=value)

    wb.save(prepared_excel)
    wb.close()

    # save trace artifacts for reproducibility
    shutil.copy2(profile_file, workspace / "profile_snapshot.yaml")
    meta_df.to_csv(csv_dir / "_meta.csv", index=False, encoding="utf-8-sig")
    styles_df.to_csv(csv_dir / "_styles.csv", index=False, encoding="utf-8-sig")
    manual_df.to_csv(csv_dir / "_manual_inputs.csv", index=False, encoding="utf-8-sig")
    (logs_dir / "source.txt").write_text(str(src), encoding="utf-8")

    return {"workspace": workspace, "prepared_excel": prepared_excel}


def parse_args(argv: list[str] | None = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Prepare workspace from source Excel/profile.")
    parser.add_argument("--excel", required=True, help="Source Excel path.")
    parser.add_argument("--profile", required=True, help="Profile YAML path.")
    parser.add_argument("--workspace-root", default="Data/intermediate/workspaces")
    parser.add_argument("--run-id", default=None)
    return parser.parse_args(argv)


if __name__ == "__main__":
    args = parse_args()
    result = prepare_workspace(
        excel_path=args.excel,
        profile_path=args.profile,
        workspace_root=args.workspace_root,
        run_id=args.run_id,
    )
    print(f"Workspace: {result['workspace']}")
    print(f"Prepared Excel: {result['prepared_excel']}")
